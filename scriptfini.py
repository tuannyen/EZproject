from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import pandas as pd
import time
import datetime
import os
import sys
import calendar
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

#Fonction pour générer le numéro de facture
def Commandetofacture(numero_commande, ieme_commande):
    char = numero_commande.replace("-A", "")  # Retirer le "-A"
    return f"FACV_{char}-{ieme_commande}"

#Fonction pour générer le numéro de commande
def FacturetoCommande(reference):
    numero_complet = reference.split("_")[1]  # Récupère la partie après "_"
    numero_final = "-".join(numero_complet.split("-")[:-1])  # Supprime la dernière partie
    return numero_final + "-A"  # Ajoute "-A"

def extraire_dernier_numero(reference):
    return reference.split("-")[-1]  # Prend la dernière partie après le dernier "-"





#################################################################################################################################
###################################### OUVRIR FICHIER EXCEL RECUPERER NUMERO COMMANDE
#################################################################################################################################
# 📌 Demande à l'utilisateur de saisir le mois et l'année
mois = input("📅 Entrez le mois (01 à 12) : ").strip()
annee = input("📅 Entrez l'année (YYYY) : ").strip()

# 📁 Nom du fichier Excel basé sur le mois et l'année choisis
EXCEL_FILE = f"TABLEAU VENTE-POUR-DE-BON-{mois}-{annee}.xlsx"

# 🔍 Lire le fichier en ignorant les 7 premières lignes
try:
    df_old = pd.read_excel(EXCEL_FILE, sheet_name="matrice", skiprows=7)  # "matrice" = le bon onglet

    # 📌 Afficher les nouvelles colonnes pour vérifier que c'est bon
    print("📌 Colonnes détectées après correction :")
    print(df_old.columns.tolist())

    # 📌 Vérifier si la colonne "Numéro de facture" est bien détectée
    if "Numéro de facture" in df_old.columns:
        factures_existantes = df_old["Numéro de facture"].dropna().astype(str).tolist()
        print(f"📊 Nombre de factures existantes : {len(factures_existantes)}")
        print(f"📄 Liste des premières factures : {factures_existantes[:5]}")
        if factures_existantes:  # Vérifie si la liste n'est pas vide
            derniere_facture = factures_existantes[-1]
        else:
            derniere_facture = None  # Ou une valeur par défaut comme "Aucune facture"

        print("✅ La derniere facture non vide est :", derniere_facture)
    else:
        print("⚠️ La colonne 'Numéro de facture' n'existe toujours pas, vérifie son orthographe.")
        factures_existantes = []

except Exception as e:
    print(f"❌ Erreur lors de la lecture du fichier : {e}")
    factures_existantes = []

if derniere_facture is not None:
    derniere_commande_excel = FacturetoCommande(derniere_facture)
    dernier_num_commande = extraire_dernier_numero(derniere_facture)
else:
    derniere_commande_excel = None  # Ou une valeur par défaut
    dernier_num_commande = dernier_num_commande = input("🔢 Entrez le dernier numéro de commande : ")  # l'utilisateur rentre le dernier numero de commande

#################################################################################################################################
################################################################ RECUPERATION DES DONNEES SUR LE SITE
#################################################################################################################################

#CONFIGURATION SELENIUM (Ouvre Chrome)
options = webdriver.ChromeOptions()
options.add_argument(r"user-data-dir=C:\Users\tuand\AppData\Local\Google\Chrome\User Data")  # Chemin du dossier User Data
options.add_argument(r"profile-directory=Default")  # Ton profil actuel
driver = webdriver.Chrome(options=options)
options.add_argument("--disable-features=IsolateOrigins,site-per-process")
options.add_argument("--disable-popup-blocking")
options.add_argument("--disable-extensions")
#options.add_argument("--headless")  # 🚀 Mode invisible (optionnel)

time.sleep(2)

# NAVIGUER VERS LA PAGE DES COMMANDES
#Calcul du premier jour du mois sélectionné (00:00:00)
date_debut = datetime.datetime(int(annee), int(mois), 1, 0, 0, 0)
#Calcul du premier jour du mois suivant (00:00:00)
if int(mois) == 12:  # Si décembre, passer à janvier de l'année suivante
    mois_suivant = 1
    annee_suivante = int(annee) + 1
else:
    mois_suivant = int(mois) + 1
    annee_suivante = int(annee)
date_fin = datetime.datetime(annee_suivante, mois_suivant, 1, 0, 0, 0)  # 1er jour du mois suivant à 00:00:00
# 📌 Conversion en timestamp Unix (millisecondes comme utilisé dans l'URL)
startDate = int(time.mktime(date_debut.timetuple()) * 1000)
endDate = int(time.mktime(date_fin.timetuple()) * 1000)  # Premier instant du mois suivant

# ✅ Générer l'URL filtrée avec le bon `endDate`
URL_COMMANDES = f"https://vendeur.pourdebon.com/mmp/shop/order/all?period=%7B%22startDate%22%3A{startDate}%2C%22endDate%22%3A{endDate}%7D&periodAuto=true&select-search=orderId&limit=200&sort=order-list-date-created-id%2CDESC&statuses=%5B%22Re%C3%A7ue%3D%3D%5C%22RECEIVED%5C%22%22%5D"
driver.get(URL_COMMANDES)

time.sleep(4)

commande_element = driver.find_elements(By.XPATH, "(//td[contains(@class, '1btx70q') and contains(@id, 'order-list-date-created-id')])[1]")[0]
hauteur_element = driver.execute_script("return arguments[0].offsetHeight;", commande_element)
print(f"📏 Hauteur détectée : {hauteur_element}px")  # Vérifier la hauteur en console
time.sleep(1)

# Liste temporaire pour stocker les nouvelles commandes avec noms des clients et dates
nouvelles_factures = []

# 🔍 Récupérer tous les numéros de commande sur la page principale (du plus récent au plus ancien)
commandes_count = len(driver.find_elements(By.XPATH, "//div[contains(@class, 'ffbKvt')]"))
print(f"📊 Nombre de commandes trouvées : {commandes_count}")
num_commande=str(int(int(dernier_num_commande)+commandes_count+1))
scroll_value=0
for i in range(commandes_count):
    # 🔄 Rafraîchir la liste des commandes après chaque navigation
    commandes_elements = driver.find_elements(By.XPATH, "//div[contains(@class, 'ffbKvt')]")

    # Vérifier qu'on ne dépasse pas la liste
    if i >= len(commandes_elements):
        break

    commande_element = commandes_elements[i]
    numero_commande = commande_element.text.strip()

    # Vérifier si on a atteint la dernière commande enregistrée
    if derniere_commande_excel is not None and numero_commande == derniere_commande_excel: 
        print(f"🚫 Commande {numero_commande} déjà enregistrée, arrêt de la récupération.")
        break  # On s'arrête dès qu'on trouve la dernière enregistrée

    # 🔍 Cliquer sur la commande pour ouvrir la page détaillée
    
    commande_element.click()
    time.sleep(3)  # Attendre le chargement de la page

    # 📌 Récupérer le nom du client
    try:
        nom_client_element = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[2]/div/div[2]/div/div/div/div[1]/div[2]/main/div/div[3]/div[2]/section/div[2]/div/p/span/span")
        nom_client = nom_client_element.text.strip()
    except:
        nom_client = "Nom introuvable"

    # 📌 Récupérer la date de commande
    try:
        date_element = driver.find_element(By.XPATH, "/html/body/div[1]/div/div[2]/div/div[2]/div/div/div/div[1]/div[2]/div/div[1]/div/p")
        date_text = date_element.text.strip()

        # Extraire uniquement la date (format JJ/MM/AAAA)
        match = re.search(r"\d{2}/\d{2}/\d{4}", date_text)
        date_commande = match.group(0) if match else "Date introuvable"
    except:
        date_commande = "Date introuvable"
    # ✅ Récupérer le Prix TTC
    prix_ttc_element = driver.find_element(By.XPATH, "(//h3[contains(@class, 'iXUXQP') and contains(text(), '€')])[1]")
    prix_ttc_text = prix_ttc_element.text.strip().replace("€", "").replace(",", ".")
    prix_ttc = float(prix_ttc_text)

    # ✅ Récupérer les prix HT à 20%
    prix_ht_elements = driver.find_elements(By.XPATH, "//div[contains(@class, 'eHcEem') and .//h3[contains(text(), 'SKU')]]//*[contains(text(), 'Prix produit total HT')]/following-sibling::*")
    tva_elements = driver.find_elements(By.XPATH, "//div[contains(@class, 'eHcEem') and .//h3[contains(text(), 'SKU')]]//div[contains(text(), 'Taxe')]")

    total_ht_20 = 0  # Initialisation du total HT 20%
    for i in range(len(prix_ht_elements)):
        prix_ht_text = prix_ht_elements[i].text.strip()
        tva_text = tva_elements[i].text.strip()

        try:
            prix_ht = float(prix_ht_text.split("x")[-1].replace("€", "").replace(",", ".").strip())
        except:
            continue  # Ignore si erreur

        if "20" in tva_text:
            total_ht_20 += prix_ht

    # Ajouter les informations à la liste
    num_commande=str(int(num_commande)-1)   
    tempNumFacture=Commandetofacture(numero_commande,num_commande)
    nouvelles_factures.append({
        "Numéro de Facture": tempNumFacture,
        "Nom du client": nom_client,
        "Date de commande": date_commande,
        "TTC" : prix_ttc,
        "Total HT 20%": total_ht_20
    })

    # 🔙 Revenir à la page principale et rafraîchir la liste
    driver.back()
    # 🔄 Attendre un peu après être revenu sur la page principale
    time.sleep(2)

    # 🔽 Scroller progressivement (augmenter de 100px à chaque itération)
    scroll_value += hauteur_element
    driver.execute_script(f"window.scrollBy(0, {scroll_value});")
      
    time.sleep(2)

# ✅ Fermer Selenium après l'extraction
driver.quit()

# 📊 Créer un DataFrame pandas pour afficher les nouvelles commandes avec les clients
df_nouvelles_factures = pd.DataFrame(nouvelles_factures)

# 📌 Afficher le tableau pour vérification
print("\n📊 Nouvelles commandes avec clients et dates :")
print(df_nouvelles_factures.to_string(index=False))  # Affichage formaté sans index













# 📌 Ajouter les nouvelles factures en respectant la structure du fichier
if nouvelles_factures:
    wb = load_workbook(EXCEL_FILE)
    ws = wb["matrice"]  # Adapter si le nom de l’onglet est différent

    # 🔍 Trouver la première ligne vide et non fusionnée
    col_num_facture = 2  # Colonne B
    first_empty_row = None

    for row in range(9, ws.max_row + 1):  # On commence après l'entête (ligne 9)
        cell = ws.cell(row=row, column=col_num_facture)

        # Vérifier si la cellule fait partie d'une cellule fusionnée
        cell_coord = f"{get_column_letter(col_num_facture)}{row}"  # Ex: B9, B10...
        is_merged = any(cell_coord in merged for merged in ws.merged_cells)

        # Si la cellule est vide et non fusionnée, on peut l'utiliser
        if cell.value is None and not is_merged:
            first_empty_row = row
            break

    if first_empty_row:
        print(f"📌 Première ligne vide et non fusionnée détectée : {first_empty_row}")

        # 🔄 Ajouter les nouvelles factures à partir de cette ligne
        row_index = first_empty_row

        for facture in reversed(nouvelles_factures):
            ws.cell(row=row_index, column=2, value=facture["Numéro de Facture"])  # Colonne B : Numéro de facture
            ws.cell(row=row_index, column=3, value=facture["Nom du client"])  # Colonne C : Nom du client
            ws.cell(row=row_index, column=1, value=facture["Date de commande"])  # Colonne A : date de commande
            ws.cell(row=row_index, column=8, value=facture["TTC"])  # Colonne H : TTC 
            ws.cell(row=row_index, column=6, value=facture["Total HT 20%"])  # Colonne F : TTHT 20%
            row_index += 1  # Passer à la ligne suivante

        # 💾 Sauvegarder sans modifier la mise en forme
        wb.save(EXCEL_FILE)
        wb.close()
        print("✅ Nouvelles factures ajoutées avec succès ! 🎉")

    else:
        print("⚠️ Aucune ligne vide et non fusionnée trouvée pour ajouter les nouvelles factures.")

else:
    print("✅ Aucune nouvelle facture à ajouter.")


